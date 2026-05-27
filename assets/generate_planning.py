"""
Génère le planning Excel du projet machine de dépose de pâte thermique.
Exécuter avec : python3 generate_planning.py
Sortie : planning.xlsx dans le même dossier.
"""

from datetime import date, timedelta
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Couleurs (ARGB)
# ---------------------------------------------------------------------------
C_HEADER_BG    = "FF1F3864"   # Bleu foncé — en-têtes
C_HEADER_FG    = "FFFFFFFF"   # Blanc
C_VACATION     = "FFFF6B35"   # Orange vif — vacances
C_MILESTONE    = "FFFFD700"   # Or — jalons
C_SOFT_GEEETECH= "FF4472C4"   # Bleu — phases logiciel Geeetech
C_HARD_CNC     = "FFE07B39"   # Orange — assemblage CNC (hardware)
C_MIGRATION    = "FF70AD47"   # Vert — phases migration/validation CNC
C_REPORT       = "FF9DC3E6"   # Bleu clair — rédaction rapport
C_BUGS         = "FFFF0000"   # Rouge — corrections bugs
C_WEEKEND      = "FFFFE699"   # Jaune pâle — week-end travaillé
C_LIGHT_GRAY   = "FFF2F2F2"   # Gris très clair — alternance lignes
C_BORDER       = "FFB8B8B8"

# ---------------------------------------------------------------------------
# Définition des phases
# ---------------------------------------------------------------------------
# type: S=Soft Geeetech, H=Hardware CNC, M=Migration/Validation CNC,
#       B=Bug fixes, R=Rapport, V=Vacances
PHASES = [
    # (id, label, type, sessions_ou_duree, date_debut, date_fin)
    ("P0",  "Phase 0 — Identification matériel",              "S", "1 session",   date(2026, 5, 27), date(2026, 5, 29)),
    ("P1",  "Phase 1 — Caméra de base",                       "S", "1 session",   date(2026, 5, 29), date(2026, 6, 2)),
    ("P2",  "Phase 2 — ArUco & calibrage",                    "S", "3 sessions",  date(2026, 6, 2),  date(2026, 6, 9)),
    ("P3",  "Phase 3 — G-code Marlin",                        "S", "2 sessions",  date(2026, 6, 9),  date(2026, 6, 13)),
    ("P4",  "Phase 4 — Interface graphique",                  "S", "3 sessions",  date(2026, 6, 10), date(2026, 6, 15)),
    # Rédaction rapport en parallèle — toute la durée du projet
    ("RD",  "Rédaction rapport (soir ~1h/j, en parallèle)",   "R", "continu",     date(2026, 5, 27), date(2026, 8, 24)),
    ("JD",  "JALON — DRAFT rapport à remettre",               "J", "",            date(2026, 6, 15), date(2026, 6, 16)),
    ("VAC", "VACANCES",                                       "V", "",            date(2026, 6, 15), date(2026, 6, 20)),
    ("P5",  "Phase 5 — Zone & trajectoire",                   "S", "3 sessions",  date(2026, 6, 22), date(2026, 6, 28)),
    ("P6",  "Phase 6 — Intégration workflow",                 "S", "3 sessions",  date(2026, 6, 25), date(2026, 7, 4)),
    ("P7",  "Phase 7 — Rapport PDF",                          "S", "2 sessions",  date(2026, 7, 1),  date(2026, 7, 6)),
    ("P8",  "Phase 8 — Tests & finitions (Geeetech)",         "S", "3 sessions",  date(2026, 7, 6),  date(2026, 7, 12)),
    ("J1",  "JALON A — Logiciel validé sur Geeetech",         "J", "",            date(2026, 7, 12), date(2026, 7, 13)),
    ("P9",  "Phase 9 — Assemblage CNC cible (hardware)",      "H", "~2 semaines", date(2026, 7, 13), date(2026, 7, 26)),
    ("P10", "Phase 10 — Portage logiciel sur CNC",            "M", "2 sessions",  date(2026, 7, 22), date(2026, 7, 26)),
    ("P11", "Phase 11 — Validation complète sur CNC",         "M", "3 sessions",  date(2026, 7, 24), date(2026, 7, 31)),
    ("J2",  "JALON B — SOUTENANCE BLANCHE",                   "J", "",            date(2026, 7, 31), date(2026, 8, 1)),
    ("P12", "Phase 12 — Corrections de bugs",                 "B", "~1 semaine",  date(2026, 8, 3),  date(2026, 8, 10)),
    ("P13", "Phase 13 — Finalisation rapport",                "R", "~3 semaines", date(2026, 8, 3),  date(2026, 8, 24)),
    ("J3",  "JALON C — SOUTENANCE FINALE",                    "J", "",            date(2026, 8, 28), date(2026, 8, 29)),
]

PHASE_COLORS = {
    "S": C_SOFT_GEEETECH,
    "H": C_HARD_CNC,
    "M": C_MIGRATION,
    "B": C_BUGS,
    "R": C_REPORT,
    "V": C_VACATION,
    "J": C_MILESTONE,
}

# ---------------------------------------------------------------------------
# Génération des semaines (lundi → dimanche)
# ---------------------------------------------------------------------------
START_DATE = date(2026, 5, 25)   # Lundi semaine 22
END_DATE   = date(2026, 9, 6)    # Après la soutenance finale

def get_weeks(start: date, end: date):
    """Retourne la liste des lundis de chaque semaine entre start et end."""
    weeks = []
    d = start - timedelta(days=start.weekday())   # Aligner sur lundi
    while d <= end:
        weeks.append(d)
        d += timedelta(weeks=1)
    return weeks

WEEKS = get_weeks(START_DATE, END_DATE)

# ---------------------------------------------------------------------------
# Helpers styles
# ---------------------------------------------------------------------------
def make_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border() -> Border:
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def medium_border() -> Border:
    s = Side(style="medium", color="FF000000")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_to_range(ws, row_min, col_min, row_max, col_max, fill=None, font=None, alignment=None, border=None):
    for r in range(row_min, row_max + 1):
        for c in range(col_min, col_max + 1):
            cell = ws.cell(row=r, column=c)
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border

# ---------------------------------------------------------------------------
# Feuille 1 — Gantt par semaine
# ---------------------------------------------------------------------------
def build_gantt(wb: openpyxl.Workbook):
    ws = wb.active
    ws.title = "Planning Gantt"
    ws.sheet_view.showGridLines = False

    # ---- Dimensions colonnes ----
    ws.column_dimensions["A"].width = 6    # ID phase
    ws.column_dimensions["B"].width = 38   # Nom phase
    ws.column_dimensions["C"].width = 13   # Sessions/durée
    # Colonnes semaines
    COL_WEEKS_START = 4
    for i, _ in enumerate(WEEKS):
        col = COL_WEEKS_START + i
        ws.column_dimensions[get_column_letter(col)].width = 8

    # ---- Ligne 1 : titre ----
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=COL_WEEKS_START + len(WEEKS) - 1)
    title_cell = ws.cell(row=1, column=1,
        value="PLANNING PROJET — Machine de Dépose de Pâte Thermique")
    title_cell.fill = make_fill(C_HEADER_BG)
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=C_HEADER_FG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Ligne 2 : sous-titre ----
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=COL_WEEKS_START + len(WEEKS) - 1)
    sub = ws.cell(row=2, column=1,
        value="BUT3 Informatique — Stage 2026  |  Vacances : 15–19 juin  |  Soutenance blanche : fin juillet  |  Soutenance finale : fin août")
    sub.fill = make_fill("FF2E4D7B")
    sub.font = Font(name="Calibri", size=10, color=C_HEADER_FG, italic=True)
    sub.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Ligne 3 : en-têtes colonnes fixes ----
    ws.row_dimensions[3].height = 20
    for col, label in [(1, "ID"), (2, "Phase"), (3, "Charge")]:
        c = ws.cell(row=3, column=col, value=label)
        c.fill = make_fill(C_HEADER_BG)
        c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    # ---- Lignes 3-4 : en-têtes semaines (mois + numéro semaine) ----
    cur_month = None
    month_start_col = COL_WEEKS_START
    month_label = ""

    for i, monday in enumerate(WEEKS):
        col = COL_WEEKS_START + i
        # Ligne 3 : numéro de semaine
        cell3 = ws.cell(row=3, column=col)
        cell3.value = monday.strftime("S%V\n%d/%m")
        cell3.fill = make_fill(C_HEADER_BG)
        cell3.font = Font(name="Calibri", bold=True, size=8, color=C_HEADER_FG)
        cell3.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell3.border = thin_border()
        ws.row_dimensions[3].height = 30

    # ---- Légende (ligne 4) ----
    ws.row_dimensions[4].height = 16
    legend_items = [
        (C_SOFT_GEEETECH, "Logiciel (Geeetech)"),
        (C_HARD_CNC,      "Hardware CNC"),
        (C_MIGRATION,     "Migration/Validation CNC"),
        (C_BUGS,          "Corrections bugs"),
        (C_REPORT,        "Rédaction rapport"),
        (C_VACATION,      "Vacances"),
        (C_MILESTONE,     "Jalons"),
    ]
    leg_col = 1
    for color, label in legend_items:
        c = ws.cell(row=4, column=leg_col, value="  " + label)
        c.fill = make_fill(color)
        c.font = Font(name="Calibri", size=8, bold=True,
                      color=C_HEADER_FG if color not in (C_MILESTONE, C_WEEKEND) else "FF000000")
        c.alignment = Alignment(vertical="center")
        c.border = thin_border()
        leg_col += 1
    # Fusionner le reste de la ligne 4
    if leg_col <= COL_WEEKS_START + len(WEEKS) - 1:
        ws.merge_cells(start_row=4, start_column=leg_col,
                       end_row=4, end_column=COL_WEEKS_START + len(WEEKS) - 1)

    # ---- Lignes de phases ----
    DATA_ROW_START = 5
    for row_idx, (pid, label, ptype, charge, d_start, d_end) in enumerate(PHASES):
        row = DATA_ROW_START + row_idx
        ws.row_dimensions[row].height = 22

        is_jalon  = (ptype == "J")
        is_vacances = (ptype == "V")
        base_fill = make_fill(C_LIGHT_GRAY if row_idx % 2 == 0 else "FFFFFFFF")
        row_fill  = make_fill(C_MILESTONE) if is_jalon else (
                    make_fill(C_VACATION) if is_vacances else base_fill)
        row_font  = Font(name="Calibri", size=10,
                         bold=is_jalon or is_vacances,
                         color="FF000000")

        # Colonne A — ID
        ca = ws.cell(row=row, column=1, value=pid)
        ca.fill = row_fill
        ca.font = Font(name="Calibri", size=9, bold=True, color="FF000000")
        ca.alignment = Alignment(horizontal="center", vertical="center")
        ca.border = thin_border()

        # Colonne B — Label
        cb = ws.cell(row=row, column=2, value=label)
        cb.fill = row_fill
        cb.font = row_font
        cb.alignment = Alignment(vertical="center")
        cb.border = thin_border()

        # Colonne C — Charge
        cc = ws.cell(row=row, column=3, value=charge)
        cc.fill = row_fill
        cc.font = Font(name="Calibri", size=9, color="FF404040", italic=True)
        cc.alignment = Alignment(horizontal="center", vertical="center")
        cc.border = thin_border()

        # Colonnes semaines — colorier si la phase est active
        phase_fill = make_fill(PHASE_COLORS.get(ptype, "FFCCCCCC"))
        for i, monday in enumerate(WEEKS):
            sunday = monday + timedelta(days=6)
            col = COL_WEEKS_START + i
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border()

            # La semaine est dans la plage de la phase si elle se chevauche
            week_in_phase = (monday < d_end) and (sunday >= d_start)
            if week_in_phase:
                cell.fill = phase_fill
                if is_jalon:
                    cell.value = "★"
                    cell.font = Font(name="Calibri", size=12, bold=True, color="FF7B4D00")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = base_fill

    # ---- Marquer les colonnes vacances ----
    vac_start = date(2026, 6, 15)
    vac_end   = date(2026, 6, 19)
    for i, monday in enumerate(WEEKS):
        sunday = monday + timedelta(days=6)
        if (monday <= vac_end) and (sunday >= vac_start):
            col = COL_WEEKS_START + i
            # Ajouter une bordure épaisse pour signaler la semaine
            for r in range(3, DATA_ROW_START + len(PHASES)):
                cell = ws.cell(row=r, column=col)
                existing_fill = cell.fill.fgColor.rgb if cell.fill.fill_type == "solid" else None
                if existing_fill is None or existing_fill == C_LIGHT_GRAY or existing_fill == "FFFFFFFF":
                    cell.fill = make_fill("FFFCE4D6")  # Fond orange très pâle

    # Figer les 3 premières colonnes + les 3 premières lignes
    ws.freeze_panes = ws.cell(row=DATA_ROW_START, column=COL_WEEKS_START)

    # ---- Ligne 3 : hauteur ajustée ----
    ws.row_dimensions[3].height = 32

# ---------------------------------------------------------------------------
# Feuille 2 — Tableau de bord sessions
# ---------------------------------------------------------------------------
def build_dashboard(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Détail phases")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    # Titre
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "DÉTAIL DES PHASES — Planning machine de dépose de pâte thermique"
    t.fill = make_fill(C_HEADER_BG)
    t.font = Font(name="Calibri", bold=True, size=13, color=C_HEADER_FG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # En-têtes colonnes
    headers = ["ID", "Phase", "Type", "Sessions", "Durée (h)", "Début", "Fin"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = make_fill(C_HEADER_BG)
        c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 20

    type_labels = {
        "S": "Logiciel (Geeetech)",
        "H": "Hardware CNC",
        "M": "Migration / CNC",
        "B": "Corrections bugs",
        "R": "Rapport",
        "V": "Vacances",
        "J": "Jalon",
    }

    session_map = {
        "P0": (1, 2),  "P1": (1, 2),  "P2": (3, 6),  "P3": (2, 4),
        "P4": (3, 6),  "P5": (3, 6),  "P6": (3, 6),  "P7": (2, 4),
        "P8": (3, 6),  "P9": (0, 0),  "P10": (2, 4), "P11": (3, 6),
        "P12": (0, 0), "P13": (0, 0), "RD": (0, 0),  "JD": (0, 0),
    }

    cumul_h = 0
    for row_idx, (pid, label, ptype, charge, d_start, d_end) in enumerate(PHASES):
        row = row_idx + 3
        ws.row_dimensions[row].height = 20
        is_jalon = (ptype == "J")
        is_vac   = (ptype == "V")

        bg = PHASE_COLORS.get(ptype, "FFCCCCCC")
        fill = make_fill(bg)
        bold = is_jalon or is_vac

        # Sessions et heures
        if pid in session_map:
            sess, hrs = session_map[pid]
            sess_str = f"{sess} session{'s' if sess > 1 else ''}" if sess else charge
            hrs_str  = f"~{hrs}h" if hrs else charge
            cumul_h += hrs
        else:
            sess_str = charge
            hrs_str  = ""

        vals = [pid, label, type_labels.get(ptype, ptype), sess_str, hrs_str,
                d_start.strftime("%d/%m/%Y"), d_end.strftime("%d/%m/%Y")]

        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.font = Font(name="Calibri", size=10, bold=bold,
                          color="FF000000" if ptype in ("J", "V") else "FF1A1A1A")
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (1, 3, 4, 5, 6, 7) else "left")
            c.border = thin_border()

    # Total
    total_row = len(PHASES) + 3
    ws.row_dimensions[total_row].height = 22
    ws.merge_cells(start_row=total_row, start_column=1,
                   end_row=total_row, end_column=4)
    tc = ws.cell(row=total_row, column=1, value="TOTAL — sessions de développement logiciel")
    tc.fill = make_fill(C_HEADER_BG)
    tc.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    th = ws.cell(row=total_row, column=5, value=f"~{cumul_h}h")
    th.fill = make_fill(C_HEADER_BG)
    th.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
    th.alignment = Alignment(horizontal="center", vertical="center")

    for col in (6, 7):
        ws.cell(row=total_row, column=col).fill = make_fill(C_HEADER_BG)

    # ---- Bloc jalons récapitulatif ----
    JAL_ROW = total_row + 2
    ws.row_dimensions[JAL_ROW].height = 20
    ws.merge_cells(start_row=JAL_ROW, start_column=1,
                   end_row=JAL_ROW, end_column=7)
    jt = ws.cell(row=JAL_ROW, column=1, value="JALONS CLÉS")
    jt.fill = make_fill(C_MILESTONE)
    jt.font = Font(name="Calibri", bold=True, size=11, color="FF7B4D00")
    jt.alignment = Alignment(horizontal="center", vertical="center")
    jt.border = medium_border()

    jalons = [
        ("JD", "Premier draft rapport à remettre",             "15 juin 2026",
         "Draft complet avant départ en vacances"),
        ("J1", "Logiciel validé sur Geeetech (PoC)",           "~12 juillet 2026",
         "Fin développement Partie A"),
        ("J2", "Soutenance blanche",                            "Fin juillet 2026",
         "CNC validée, démo complète"),
        ("J3", "Soutenance finale",                             "Fin août 2026",
         "Rapport remis + présentation"),
    ]
    jal_headers = ["ID", "Jalon", "Date cible", "Critère de succès", "", "", ""]
    for col, h in enumerate(jal_headers[:4], start=1):
        c = ws.cell(row=JAL_ROW + 1, column=col, value=h)
        c.fill = make_fill("FF7B4D00")
        c.font = Font(name="Calibri", bold=True, size=9, color=C_HEADER_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    for col in range(5, 8):
        c = ws.cell(row=JAL_ROW + 1, column=col)
        c.fill = make_fill("FF7B4D00")

    for j_idx, (jid, jlabel, jdate, jcrit) in enumerate(jalons):
        jrow = JAL_ROW + 2 + j_idx
        ws.row_dimensions[jrow].height = 20
        # Colorier en rouge vif si c'est le draft (jalon critique)
        bg = "FFFFE4B5" if jid != "JD" else "FFFF6B6B"
        ws.merge_cells(start_row=jrow, start_column=4, end_row=jrow, end_column=7)
        for col, val in enumerate([jid, jlabel, jdate, jcrit], start=1):
            c = ws.cell(row=jrow, column=col, value=val)
            c.fill = make_fill(bg)
            c.font = Font(name="Calibri", size=10, bold=(col == 1 or jid == "JD"),
                          color="FF000000")
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col in (1, 3) else "left")
            c.border = thin_border()

# ---------------------------------------------------------------------------
# Feuille 3 — Calendrier sessions week-end
# ---------------------------------------------------------------------------
def build_weekend_calendar(wb: openpyxl.Workbook):
    ws = wb.create_sheet("Sessions week-end")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40

    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "SESSIONS WEEK-END DISPONIBLES (2–3h chacune)"
    t.fill = make_fill(C_HEADER_BG)
    t.font = Font(name="Calibri", bold=True, size=12, color=C_HEADER_FG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col, h in enumerate(["Samedi", "Dimanche", "Phase recommandée", "Objectif de session"], 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill = make_fill(C_HEADER_BG)
        c.font = Font(name="Calibri", bold=True, size=10, color=C_HEADER_FG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 18

    # Lister tous les week-ends disponibles (hors vacances)
    weekends = []
    d = date(2026, 5, 30)   # Premier samedi disponible
    while d <= date(2026, 8, 30):
        # Samedi de cette semaine
        sat = d
        sun = d + timedelta(days=1)
        # Exclure vacances (15–19 juin, vacances = dim 14 juin OK mais lun 15 → ven 19 non)
        vac_start = date(2026, 6, 15)
        vac_end   = date(2026, 6, 19)
        sat_ok = not (vac_start <= sat <= vac_end)
        sun_ok = not (vac_start <= sun <= vac_end)
        weekends.append((sat if sat_ok else None, sun if sun_ok else None))
        d += timedelta(weeks=1)

    # Phases correspondant à chaque week-end
    phase_suggestions = [
        ("Phase 1 + Rapport (soir)",  "Finir camera.py + rédiger intro rapport"),
        ("Phase 2 + Rapport (soir)",  "ArUco + rédiger section système physique"),
        ("Phase 2–3 + Rapport (soir)","Ajustements ArUco + rédiger section architecture"),
        ("Phase 3–4 + Rapport (soir)","G-code + démarrer GUI + rédiger section logicielle"),
        ("Phase 4 + DRAFT RAPPORT",   "Tests tactile Rpi + finaliser draft avant le 15/06"),
        ("VACANCES — repos",          ""),
        ("Phase 5–6 + Rapport",       "Widget zone + intégration + enrichir rapport"),
        ("Phase 6 + Rapport",         "Cycle complet Geeetech + rédiger section résultats"),
        ("Phase 7–8 + Rapport",       "PDF + tests + photos pour le rapport"),
        ("Phase 8 + Rapport",         "Finitions Geeetech + mise à jour rapport"),
        ("Phase 9 (hardware)",        "Assemblage CNC — câblage moteurs"),
        ("Phase 9–10 + Rapport",      "Tests CNC + portage + section intégration rapport"),
        ("Phase 11 + Rapport",        "Validation CNC + rédiger section validation"),
        ("Phase 12 + Rapport",        "Corrections bugs + relecture rapport"),
        ("Phase 13 — Rapport",        "Rédaction intensive — section résultats & bilan"),
        ("Phase 13 — Rapport",        "Rédaction intensive — figures, captures d'écran"),
        ("Phase 13 — Rapport",        "Relecture complète + corrections"),
        ("Phase 13 — Rapport",        "Finalisation rapport + remise"),
    ]

    for row_idx, (sat, sun) in enumerate(weekends[:len(phase_suggestions)]):
        row = row_idx + 3
        ws.row_dimensions[row].height = 20
        sat_str = sat.strftime("%d/%m") if sat else "—  (vacances)"
        sun_str = sun.strftime("%d/%m") if sun else "—  (vacances)"
        phase_str, obj_str = phase_suggestions[row_idx]

        is_vac  = (sat is None or sun is None) or ("VACANCES" in phase_str)
        bg_hex  = C_VACATION if is_vac else (C_WEEKEND if row_idx % 2 == 0 else "FFFFFFD0")
        f = make_fill(bg_hex)
        fn = Font(name="Calibri", size=10, bold=is_vac,
                  color="FF000000")

        for col, val in enumerate([sat_str, sun_str, phase_str, obj_str], 1):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = f
            c.font = fn
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col < 3 else "left")
            c.border = thin_border()

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    wb = openpyxl.Workbook()
    build_gantt(wb)
    build_dashboard(wb)
    build_weekend_calendar(wb)

    output_path = "planning.xlsx"
    wb.save(output_path)
    print(f"Planning généré : {output_path}")

if __name__ == "__main__":
    main()
